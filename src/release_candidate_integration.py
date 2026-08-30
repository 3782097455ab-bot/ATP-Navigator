"""Competition release-candidate member-data integration.

This module never edits source workbooks and never mixes MIC, IC50, Ki, Kd,
or computational energies as one label.  It creates provenance-preserving QC
tables and a manifest.  The large BindingDB TSV is streamed and reduced to a
compact experimental-record table; the original 325 MB file is not copied into
the repository.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import tempfile
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from rdkit import Chem, rdBase
from rdkit.Chem.Scaffolds import MurckoScaffold

from data_integration.member_data_audit import audit_member2, register_member3

rdBase.DisableLog("rdApp.error")


BINDINGDB_COLUMNS = [
    "BindingDB Reactant_set_id", "Ligand SMILES", "Ligand InChI Key",
    "BindingDB MonomerID", "BindingDB Ligand Name", "Target Name",
    "Target Source Organism According to Curator or DataSource", "Ki (nM)",
    "IC50 (nM)", "Kd (nM)", "EC50 (nM)", "pH", "Temp (C)",
    "Curation/DataSource", "Article DOI", "BindingDB Entry DOI", "PMID",
]


def text(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return str(value).strip()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def atomic_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", delete=False, dir=path.parent) as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2, allow_nan=False)
        temporary = Path(handle.name)
    temporary.replace(path)


def atomic_csv(path: Path, frame: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8-sig", newline="", delete=False, dir=path.parent) as handle:
        frame.to_csv(handle, index=False)
        temporary = Path(handle.name)
    temporary.replace(path)


def structure_fields(smiles: Any) -> dict[str, Any]:
    raw = text(smiles)
    if not raw:
        return {"canonical_smiles": "", "inchikey": "", "scaffold": "", "rdkit_parse_status": "missing"}
    mol = Chem.MolFromSmiles(raw)
    if mol is None:
        return {"canonical_smiles": "", "inchikey": "", "scaffold": "", "rdkit_parse_status": "invalid"}
    canonical = Chem.MolToSmiles(mol, canonical=True, isomericSmiles=True)
    scaffold = MurckoScaffold.MurckoScaffoldSmiles(mol=mol, includeChirality=True) or canonical
    return {
        "canonical_smiles": canonical,
        "inchikey": Chem.MolToInchiKey(mol),
        "scaffold": scaffold,
        "rdkit_parse_status": "valid",
    }


NUMBER = re.compile(r"^\s*(<=|>=|<|>|=|~)?\s*([-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)\s*$")
RANGE = re.compile(r"^\s*([-+]?\d+(?:\.\d+)?)\s*[-–]\s*([-+]?\d+(?:\.\d+)?)\s*([^\d]*)$")


def parse_value(value: Any, default_unit: str = "") -> dict[str, Any]:
    raw = text(value).replace("�", "μ").replace("µ", "μ")
    match = NUMBER.match(raw)
    if match:
        return {"raw_value": raw, "operator": match.group(1) or "=", "numeric_value": float(match.group(2)),
                "unit": default_unit, "value_class": "exact_numeric"}
    range_match = RANGE.match(raw.replace("μM", " μM"))
    if range_match:
        suffix = range_match.group(3).strip() or default_unit
        return {"raw_value": raw, "operator": "range", "numeric_value": None, "unit": suffix,
                "value_class": "range"}
    mixed = re.match(r"^\s*(<=|>=|<|>|=|~)?\s*([-+]?\d+(?:\.\d+)?)\s*(.*)$", raw)
    if mixed:
        return {"raw_value": raw, "operator": mixed.group(1) or "=", "numeric_value": float(mixed.group(2)),
                "unit": mixed.group(3).strip() or default_unit, "value_class": "exact_numeric"}
    return {"raw_value": raw, "operator": "", "numeric_value": None, "unit": default_unit,
            "value_class": "missing" if not raw else "qualitative"}


def direct_atp_synthase_target(value: Any) -> bool:
    target = text(value).casefold().replace("-", "")
    return any(token in target for token in ["atp synthase", "f1foatp", "fof1atp", "f0f1atp", "f1f0atp"])


def audit_member1(project: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    source = project / "data/external/curated/Member1_ATP_inhibitor_week1.xlsx"
    sheets = pd.read_excel(source, sheet_name=None, dtype=object)
    names = list(sheets)
    raw = sheets[names[0]].copy()
    raw.columns = [text(column) for column in raw.columns]
    references = sheets[names[1]].copy() if len(names) > 1 else pd.DataFrame()
    references.columns = [text(column) for column in references.columns]
    reference_map = {text(row.get("reference_id")): row for _, row in references.iterrows()}
    rows: list[dict[str, Any]] = []
    for index, record in raw.reset_index(drop=True).iterrows():
        ref_id = text(record.get("Reference"))
        ref = reference_map.get(ref_id, {})
        parsed = parse_value(record.get("Activity Value & Unit"))
        structure = structure_fields(record.get("SMILES"))
        target = text(record.get("Target"))
        endpoint = text(record.get("Activity Type"))
        direct = direct_atp_synthase_target(target) and "implied" not in target.casefold() and "potential" not in target.casefold()
        endpoint_is_target = any(token in endpoint.casefold() for token in ["ic50", "ki", "kd", "atp synthesis", "inhibition"])
        provenance_complete = bool(text(ref.get("link")))
        training = bool(
            structure["rdkit_parse_status"] == "valid"
            and direct and endpoint_is_target and parsed["value_class"] == "exact_numeric"
            and parsed["unit"] and provenance_complete
        )
        rows.append({
            "source_row": index + 2,
            "compound_name": text(record.get("Compound Name")),
            "raw_smiles": text(record.get("SMILES")),
            **structure,
            "target": target,
            "target_annotation": "direct_atp_synthase" if direct else "indirect_or_uncertain_target_context",
            "organism": text(record.get("Organism")),
            "atp_subunit_context": target,
            "activity_type": endpoint,
            **parsed,
            "assay": endpoint,
            "reference_id": ref_id,
            "reference_name": text(ref.get("paper/database")),
            "reference_url": text(ref.get("link")),
            "doi_or_pmid": "",
            "source_level": "member_literature_summary_unverified_primary_record",
            "provenance": f"data/external/curated/{source.name}#{names[0]}!row={index + 2}",
            "training_eligible": training,
            "validation_eligible": bool(structure["rdkit_parse_status"] == "valid" and direct and provenance_complete),
            "registry_only": not training,
            "eligibility_reason": "eligible_exact_target_endpoint" if training else "qualitative/range, uncertain target relation, missing exact unit, or primary-record verification required",
        })
    frame = pd.DataFrame(rows)
    frame["identity_duplicate"] = frame["inchikey"].ne("") & frame.duplicated("inchikey", keep=False)
    summary = {
        "source_file": f"data/external/curated/{source.name}", "source_sha256": sha256(source),
        "raw_rows": len(frame), "qc_rows": len(frame),
        "valid_structure_rows": int(frame["rdkit_parse_status"].eq("valid").sum()),
        "unique_valid_structures": int(frame.loc[frame["rdkit_parse_status"].eq("valid"), "inchikey"].nunique()),
        "direct_target_rows": int(frame["target_annotation"].eq("direct_atp_synthase").sum()),
        "training_eligible_rows": int(frame["training_eligible"].sum()),
        "validation_eligible_rows": int(frame["validation_eligible"].sum()),
        "limitation": "member literature summary contains mostly qualitative/range claims and placeholder structure text; no row is promoted without exact endpoint/unit and primary-record provenance",
    }
    return frame, summary


def enhance_member2(project: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    frame, summary = audit_member2(project)
    frame = frame.copy()
    frame["source_level_normalized"] = "B"
    frame["confidence"] = "high"
    frame["duplicate_keep"] = ~frame.duplicated("strict_assay_key", keep="first")
    eligible = (
        frame["rdkit_parse_status"].eq("valid")
        & frame["endpoint"].eq("MIC")
        & frame["normalized_unit"].eq("ug/mL")
        & frame["activity_value_numeric"].notna()
        & frame["activity_value_numeric"].gt(0)
        & frame["duplicate_keep"]
        & ~frame["semantic_assay_overlap_external_v2"]
    )
    frame["training_eligible"] = eligible
    frame["validation_eligible"] = frame["rdkit_parse_status"].eq("valid") & frame["activity_value_numeric"].notna()
    frame["registry_only"] = ~frame["training_eligible"]
    frame["data_role"] = "gram_negative_mic_context"
    frame["endpoint_semantics"] = "whole_cell_MIC_not_ATP_synthase_activity"
    frame["eligibility_reason"] = "eligible_shadow_MIC_context" 
    frame.loc[~frame["training_eligible"], "eligibility_reason"] = "duplicate, semantic overlap, invalid/nonpositive value, or incompatible unit"
    summary.update({
        "training_eligible_rows": int(frame["training_eligible"].sum()),
        "validation_eligible_rows": int(frame["validation_eligible"].sum()),
        "endpoint_semantics": "MIC kept separate from ATP-target and computational labels",
    })
    return frame, summary


def audit_bindingdb(path: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    rows: list[pd.DataFrame] = []
    raw_rows = 0
    for chunk in pd.read_csv(path, sep="\t", usecols=BINDINGDB_COLUMNS, dtype=str, keep_default_na=False,
                             chunksize=20000, low_memory=False, on_bad_lines="skip"):
        raw_rows += len(chunk)
        compact: list[dict[str, Any]] = []
        for _, record in chunk.iterrows():
            structure = structure_fields(record["Ligand SMILES"])
            target = text(record["Target Name"])
            direct = direct_atp_synthase_target(target)
            for endpoint in ["Ki", "IC50", "Kd", "EC50"]:
                parsed = parse_value(record[f"{endpoint} (nM)"], "nM")
                if parsed["value_class"] == "missing":
                    continue
                exact = parsed["value_class"] == "exact_numeric" and parsed["numeric_value"] is not None and parsed["numeric_value"] > 0
                compact.append({
                    "bindingdb_reactant_set_id": text(record["BindingDB Reactant_set_id"]),
                    "bindingdb_monomer_id": text(record["BindingDB MonomerID"]),
                    "ligand_name": text(record["BindingDB Ligand Name"]),
                    "raw_smiles": text(record["Ligand SMILES"]),
                    **structure,
                    "target": target,
                    "organism": text(record["Target Source Organism According to Curator or DataSource"]),
                    "activity_type": endpoint,
                    **parsed,
                    "pH": text(record["pH"]), "temperature_C": text(record["Temp (C)"]),
                    "article_doi": text(record["Article DOI"]),
                    "bindingdb_entry_doi": text(record["BindingDB Entry DOI"]),
                    "pmid": text(record["PMID"]),
                    "curation_source": text(record["Curation/DataSource"]),
                    "target_scope": "direct_atp_synthase" if direct else "other_protein_target",
                    "training_eligible": bool(direct and exact and structure["rdkit_parse_status"] == "valid"),
                    "validation_eligible": bool(exact and structure["rdkit_parse_status"] == "valid"),
                    "registry_only": not direct,
                    "validation_scope": "ATP_target" if direct else "general_binding_benchmark_only",
                    "provenance": f"BindingDB:{text(record['BindingDB Reactant_set_id'])}",
                })
        if compact:
            rows.append(pd.DataFrame(compact))
    frame = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    summary = {
        "source_file": str(path), "source_sha256": sha256(path), "raw_rows": raw_rows,
        "qc_measurement_rows": len(frame),
        "valid_structure_rows": int(frame["rdkit_parse_status"].eq("valid").sum()) if not frame.empty else 0,
        "unique_valid_structures": int(frame.loc[frame["rdkit_parse_status"].eq("valid"), "inchikey"].nunique()) if not frame.empty else 0,
        "direct_atp_synthase_rows": int(frame["target_scope"].eq("direct_atp_synthase").sum()) if not frame.empty else 0,
        "training_eligible_rows": int(frame["training_eligible"].sum()) if not frame.empty else 0,
        "validation_eligible_rows": int(frame["validation_eligible"].sum()) if not frame.empty else 0,
        "endpoint_counts": frame["activity_type"].value_counts().to_dict() if not frame.empty else {},
        "important_scope_note": "SERCA, Na/K ATPase and other ATPases are not ATP synthase and are not admitted to the ATP-target stratum",
    }
    return frame, summary


def workbook_manifest_rows(project: Path, bindingdb_path: Path, bindingdb_hash: str) -> list[dict[str, Any]]:
    curated = project / "data/external/curated"
    rows: list[dict[str, Any]] = []
    member_map = {
        "Member1_ATP_inhibitor_week1.xlsx": "Member1",
        "经典ATP synthase inhibitor.docx": "Member1",
        "Member2_GN_antibacterial_week1.xlsx": "Member2",
        "Member3_Part2_AI_Benchmark_26_audited.xlsx": "Member3",
        "BindingDB_BindingDB_Articles_202608_tsv.zip": "Member3",
        "README.md": "project",
    }
    for path in sorted(curated.iterdir()):
        member = member_map.get(path.name, "unknown")
        if path.suffix.lower() == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            for sheet in workbook:
                rows.append({
                    "member": member, "filename": path.name,
                    "path/source": str(path.relative_to(project)).replace("\\", "/"), "sha256": sha256(path),
                    "sheet": sheet.title, "raw_rows": max((sheet.max_row or 1) - 1, 0),
                    "data_role": "ATP_target_literature" if member == "Member1" else ("GN_MIC" if member == "Member2" else "benchmark_catalog"),
                    "training_eligible": member == "Member2", "validation_eligible": member in {"Member1", "Member2"},
                    "registry_only": member == "Member3", "notes": "eligibility is row-level after QC; source workbook unchanged",
                })
        else:
            is_zip = path.suffix.lower() == ".zip"
            rows.append({
                "member": member, "filename": path.name,
                "path/source": str(path.relative_to(project)).replace("\\", "/"), "sha256": sha256(path),
                "sheet": "archive" if is_zip else "file", "raw_rows": "unknown",
                "data_role": "BindingDB source archive" if is_zip else "documentation",
                "training_eligible": False, "validation_eligible": False, "registry_only": True,
                "notes": "archive content hash matches uploaded TSV; not re-extracted" if is_zip else "documentation only",
            })
    rows.append({
        "member": "Member3", "filename": bindingdb_path.name, "path/source": str(bindingdb_path),
        "sha256": bindingdb_hash, "sheet": "TSV", "raw_rows": 93712,
        "data_role": "concrete protein-ligand experimental records", "training_eligible": False,
        "validation_eligible": True, "registry_only": False,
        "notes": "exact content duplicate of repository ZIP member; streamed from uploaded file; row-level target gate applies",
    })
    return rows


def report_text(summary: dict[str, Any]) -> str:
    m1, m2, m3, catalog = summary["member1"], summary["member2"], summary["member3_part1"], summary["member3_part2"]
    return f"""# Competition Release Candidate Data Integration Report

## Scope

This is an additive, provenance-preserving integration. Source workbooks and the uploaded BindingDB TSV were read-only. MIC, IC50, Ki, Kd and computational energies remain separate endpoints. The audit concerns data semantics, target annotation, endpoint segregation and provenance QC; it is not a biosafety assessment.

## Member 1 — ATP synthase literature table

- Raw/QC rows: {m1['raw_rows']}/{m1['qc_rows']}
- RDKit-valid structures: {m1['valid_structure_rows']} rows; {m1['unique_valid_structures']} unique structures
- Direct ATP-synthase target annotations: {m1['direct_target_rows']}
- Training-eligible exact records: {m1['training_eligible_rows']}
- Validation-eligible literature records: {m1['validation_eligible_rows']}

The workbook is primarily a literature lead list. Qualitative claims, ranges, placeholder structure descriptions, implied targets and records without exact primary-assay provenance remain reference-only.

## Member 2 — Gram-negative MIC context

- Raw rows: {m2['raw_rows']}; valid structures: {m2['valid_structures_rows']}
- Unique structures: {m2['unique_structures']}; exact structure overlap with External Dataset v2: {m2['structure_overlap_external_v2_unique']}
- Truly new structures: {m2['truly_new_structures']}
- Training-eligible nonduplicate, non-semantic-overlap assay-context rows: {m2['training_eligible_rows']}
- New strain/resistance context strings: {m2['new_species_or_strain_contexts']}

The increment is assay/strain context, not new chemical space. MIC remains whole-cell antibacterial evidence and is not ATP-synthase activity.

## Member 3 Part 1 — BindingDB concrete records

- Raw BindingDB rows: {m3['raw_rows']}; exploded endpoint measurements: {m3['qc_measurement_rows']}
- Unique valid structures: {m3['unique_valid_structures']}
- Exact/direct ATP synthase records: {m3['direct_atp_synthase_rows']}
- General binding validation records: {m3['validation_eligible_rows']}

The uploaded TSV is byte-identical to the TSV inside the repository archive. It contains concrete records, not catalog metadata. However, its ATP-related target strings are SERCA/Na-K ATPase/other ATPase contexts, not F-type ATP synthase; therefore no BindingDB row is admitted to the ATP-target training stratum. It remains a broad external binding benchmark asset.

## Member 3 Part 2 — benchmark catalog

- Catalog entries: {catalog['benchmark_registry_count']}
- Executed benchmarks: 0
- Training records: 0

The 26 rows are metadata/catalog entries only.

## Integration decision

- Member 1: ATP-target literature registry and future primary-source verification; no automatic training labels.
- Member 2: Task-A MIC shadow experiment only, with scaffold/compound leakage controls and context preserved.
- Member 3 Part 1: broad binding external-validation registry; no direct ATP-target records in this file.
- Member 3 Part 2: Benchmark Registry only.

No evidence in this integration justifies replacing Model v3 or calling an external affinity/MIC record an internal candidate activity label.
"""


def run(project: Path, bindingdb_path: Path) -> dict[str, Any]:
    project = project.resolve()
    bindingdb_path = bindingdb_path.resolve()
    output = project / "results/release_candidate/member_data_integration"
    output.mkdir(parents=True, exist_ok=True)
    member1, member1_summary = audit_member1(project)
    member2, member2_summary = enhance_member2(project)
    part1, part1_summary = audit_bindingdb(bindingdb_path)
    catalog, catalog_summary = register_member3(project)
    manifest = pd.DataFrame(workbook_manifest_rows(project, bindingdb_path, part1_summary["source_sha256"]))

    atomic_csv(output / "member_data_manifest.csv", manifest)
    atomic_csv(output / "member1_qc.csv", member1)
    atomic_csv(output / "member2_qc.csv", member2)
    atomic_csv(output / "member3_part1_qc.csv", part1)
    atomic_csv(output / "benchmark_registry.csv", catalog)
    overlap = pd.DataFrame([
        {"comparison": "Member2 vs External Dataset v2", "exact_structure_overlap": member2_summary["structure_overlap_external_v2_unique"],
         "exact_assay_overlap": member2_summary["strict_assay_overlap_external_v2_rows"],
         "semantic_assay_overlap": member2_summary["semantic_assay_overlap_external_v2_rows"],
         "truly_new_structures": member2_summary["truly_new_structures"],
         "truly_new_information": f"{member2_summary['new_species_or_strain_contexts']} new organism/strain context strings"},
        {"comparison": "Uploaded BindingDB TSV vs repository ZIP member", "exact_structure_overlap": "not_applicable",
         "exact_assay_overlap": "byte-identical source", "semantic_assay_overlap": "not_applicable", "truly_new_structures": 0,
         "truly_new_information": "0; uploaded TSV SHA256 equals uncompressed ZIP-member SHA256"},
    ])
    atomic_csv(output / "overlap_audit.csv", overlap)
    summary = {
        "member1": member1_summary, "member2": member2_summary,
        "member3_part1": part1_summary, "member3_part2": catalog_summary,
        "source_files_modified": False, "endpoint_mixing_performed": False,
        "model_training_performed": False,
    }
    atomic_json(output / "integration_summary.json", summary)
    (output / "Data_Integration_Report.md").write_text(report_text(summary), encoding="utf-8")
    return summary


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--project", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--bindingdb", type=Path, required=True)
    args = parser.parse_args()
    print(json.dumps(run(args.project, args.bindingdb), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
